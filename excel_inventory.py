

''' START OF EXCEL FILE PROGRAM '''


import openpyxl

wb = openpyxl.load_workbook('C:\\Users\sadiq\Desktop\Inventory.xlsx')

sheet = wb.get_sheet_by_name('Toyes')



item_name =[]
rate = []
qty =[]
for i in range(1, 82):

    item_name = sheet.cell(row=i , column=2).value
    print("item Name =>",item_name)
    rate = sheet.cell(row=i , column=5).value
    print("Price => ",rate)
    qty = sheet.cell(row=i , column=6).value
    print("Quantities => ",qty)
    print('*******************************************')
   

''' END OF EXCELL FILE PROGRAM '''


for i in range(1,10):
    
    baconFile = open('C:\\Users\sadiq\Desktop\sadiq.txt', 'w')
    baconFile.write(item_name[i])
    baconFile.close()
    