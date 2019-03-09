
import openpyxl

wb = openpyxl.load_workbook('C:\\Users\sadiq\Desktop\Inventory.xlsx')

sheet = wb.get_sheet_by_name('Toyes')



item_name =[]
rate = []
qty =[]
while True:
    for i in range(1, 200):

    item_name = item_name +[sheet.cell(row=i , column=2).value]
  #  print("item Name =>",item_name)
    
    rate = rate+[sheet.cell(row=i , column=5).value]
   # print("Price => ",rate)
    qty = qty +[sheet.cell(row=i , column=6).value]
    #print("Quantities => ",qty)
    #print('*******************************************')


baconFile = open('C:\\Users\sadiq\Desktop\item.txt', 'w')
baconFile.write(str(item_name))
baconFile.close()

baconFile = open('C:\\Users\sadiq\Desktop\price.txt', 'w')
baconFile.write(str(rate))
baconFile.close()

baconFile = open('C:\\Users\sadiq\Desktop\quantity.txt', 'w')
baconFile.write(str(rate))
baconFile.close()


baconFile = open('C:\\Users\sadiq\Desktop\price.txt')
content = baconFile.read()
baconFile.close()
print(content)