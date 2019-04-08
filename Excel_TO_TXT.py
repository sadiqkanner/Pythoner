
import openpyxl

wb = openpyxl.load_workbook('C:\\Users\sadiq\Desktop\Inventory.xlsx')

sheet = wb.get_sheet_by_name('Birthday dec. item')



item_name =[]
rate = []
sale = []
vender = []
qty =[]
P_units = []

for i in range(1, 95):

    item_name = item_name +[sheet.cell(row=i , column=2).value]
    rate = rate+[sheet.cell(row=i , column=4).value]
    sale = sale +[sheet.cell(row=i , column=5).value]
    vender = vender +[sheet.cell(row=i , column=3).value]
    qty = qty +[sheet.cell(row=i , column=6).value]
    P_units = P_units +[sheet.cell(row=i , column=7).value]
    #print("Quantities => ",qty)
    #print('*******************************************')


baconFile = open('C:\\Users\sadiq\Desktop\Inventory\Birthday dec_item\Items.txt', 'w')
baconFile.write(str(item_name))
baconFile.close()


baconFile = open('C:\\Users\sadiq\Desktop\Inventory\Birthday dec_item\P_Rate.txt', 'w')
baconFile.write(str(rate))
baconFile.close()


baconFile = open('C:\\Users\sadiq\Desktop\Inventory\Birthday dec_item\S_Rate.txt', 'w')
baconFile.write(str(sale))
baconFile.close()


baconFile = open('C:\\Users\sadiq\Desktop\Inventory\Birthday dec_item\Vendor.txt', 'w')
baconFile.write(str(vender))
baconFile.close()


baconFile = open('C:\\Users\sadiq\Desktop\Inventory\Birthday dec_item\Quantity.txt', 'w')
baconFile.write(str(qty))
baconFile.close()



baconFile = open('C:\\Users\sadiq\Desktop\Inventory\Birthday dec_item\P_Units.txt', 'w')
baconFile.write(str(P_units))
baconFile.close()

print("Successfully Done !!!")

baconFile = open('C:\\Users\sadiq\Desktop\price.txt')
content = baconFile.read()
baconFile.close()
print(content)